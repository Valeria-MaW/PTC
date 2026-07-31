import openpyxl
import numpy as np
import torch


class UnNormalize(object):
    def __init__(self, mean, std):
        self.mean = mean
        self.std = std

    def __call__(self, image):
        image2 = torch.clone(image)
        for t, m, s in zip(image2, self.mean, self.std):
            t.mul_(s).add_(m)
        return image2
def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet['A1'].value is None:
        sheet['A1'] = 'Model'
        sheet['B1'] = 'CLIP'
        sheet['C1'] = 'VFM'
        sheet['D1'] = 'PTC'
        sheet['E1'] = 'Min_Seeds'
        sheet['F1'] = 'Mu'
        sheet['G1'] = 'Dataset'
        sheet['H1'] = 'mIoU'

    last_row = sheet.max_row

    for index, result in enumerate(experiment_data, start=1):
        sheet.cell(row=last_row + index, column=1, value=result['Model'])
        sheet.cell(row=last_row + index, column=2, value=result['CLIP'])
        sheet.cell(row=last_row + index, column=3, value=result['VFM'])
        sheet.cell(row=last_row + index, column=4, value=result['PTC'])
        sheet.cell(row=last_row + index, column=5, value=result['Min_Seeds'])
        sheet.cell(row=last_row + index, column=6, value=result['Mu'])
        sheet.cell(row=last_row + index, column=7, value=result['Dataset'])
        sheet.cell(row=last_row + index, column=8, value=result['mIoU'])

    workbook.save(file_path)


